from openpyxl import Workbook
import pandas as pd
import xlrd

df = pd.read_excel (r'..\data\FLT+PTN_31MAR2019-26OCT2019_REV_20190607.xlsx')
#for an earlier version of Excel, you may need to use the file extension of 'xls'
print (df)

loc = ("..\data\FLT+PTN_31MAR2019-26OCT2019_REV_20190607.xlsx")

wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0,0)

for i in range(sheet.ncols):
    print(sheet.cell_value(0,i))




# wb = Workbook()
#
# # grab the active worksheet
# ws = wb.active
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")